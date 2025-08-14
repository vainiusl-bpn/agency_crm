import os
from datetime import datetime, timedelta
from flask import render_template, redirect, url_for, flash, request, current_app, send_from_directory, abort, Response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from wtforms import SelectField
from wtforms.validators import DataRequired
from app.clients import bp
from openpyxl import Workbook
from io import BytesIO
from app.clients.forms import (CompanyForm, AgreementForm, BrandForm, ClientContactForm, 
                              BrandTeamForm, PlanningInfoForm, CommitmentForm, 
                              StatusUpdateForm, MediaGroupForm, KeyMeetingForm, KeyLinkForm, GiftForm,
                              TaskTemplateForm, BrandTaskForm, TaskCompletionForm, SubcompanyForm, InvoiceForm,
                              SubbrandForm, MediaPlanForm, DigitalInfoForm, DigitalInfoLinkForm)
from app.models import (Company, Agreement, Brand, ClientContact, BrandTeam, 
                       PlanningInfo, Commitment, StatusUpdate, MediaGroup, User,
                       KeyMeeting, KeyLink, PlanningAttachment, MeetingAttachment, Gift,
                       TaskTemplate, BrandTask, TaskCompletion, Invoice, InvoiceAttachment, Subbrand, MediaPlan,
                       DigitalInfo, DigitalInfoLink)
from app import db

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

@bp.route('/companies')
@login_required
def companies():
    # Only get parent companies (those without parent_company_id)
    companies = Company.query.filter_by(parent_company_id=None).order_by(Company.name).all()
    return render_template('clients/companies.html', companies=companies)

@bp.route('/company/new', methods=['GET', 'POST'])
@login_required
def new_company():
    form = CompanyForm()
    # Get list of companies for parent company selection
    form.parent_company_id.choices = [(0, 'None')] + [(c.id, c.name) for c in Company.query.order_by(Company.name).all()]
    
    if form.validate_on_submit():
        company = Company(
            name=form.name.data,
            vat_code=form.vat_code.data,
            registration_number=form.registration_number.data,
            address=form.address.data,
            bank_account=form.bank_account.data,
            agency_fees=form.agency_fees.data,
            parent_company_id=form.parent_company_id.data if form.parent_company_id.data != 0 else None,
            status=form.status.data
        )
        db.session.add(company)
        db.session.commit()
        flash('Company created successfully!', 'success')
        return redirect(url_for('clients.company_detail', company_id=company.id))
    return render_template('clients/company_form.html', form=form, title='New Company')

@bp.route('/company/<int:company_id>')
@login_required
def company_detail(company_id):
    company = Company.query.get_or_404(company_id)
    return render_template('clients/company_detail.html', company=company, datetime=datetime)

@bp.route('/company/<int:company_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_company(company_id):
    company = Company.query.get_or_404(company_id)
    form = CompanyForm(company=company)
    
    # Get list of companies for parent company selection, excluding current company and its subcompanies
    exclude_ids = [company.id] + [s.id for s in company.subcompanies]
    parent_choices = [(0, 'None')]
    parent_choices += [(c.id, c.name) for c in Company.query.filter(~Company.id.in_(exclude_ids)).order_by(Company.name).all()]
    form.parent_company_id.choices = parent_choices
    
    if form.validate_on_submit():
        company.name = form.name.data
        company.vat_code = form.vat_code.data
        company.registration_number = form.registration_number.data
        company.address = form.address.data
        company.bank_account = form.bank_account.data
        company.agency_fees = form.agency_fees.data
        company.parent_company_id = form.parent_company_id.data if form.parent_company_id.data != 0 else None
        company.status = form.status.data
        company.updated_at = datetime.utcnow()
        db.session.commit()
        flash('Company updated successfully!', 'success')
        return redirect(url_for('clients.company_detail', company_id=company.id))
    
    elif request.method == 'GET':
        form.name.data = company.name
        form.vat_code.data = company.vat_code
        form.registration_number.data = company.registration_number
        form.address.data = company.address
        form.bank_account.data = company.bank_account
        form.agency_fees.data = company.agency_fees
        form.parent_company_id.data = company.parent_company_id or 0
        form.status.data = company.status
    
    return render_template('clients/company_form.html', form=form, title='Edit Company', company=company)

@bp.route('/company/<int:company_id>/delete', methods=['POST'])
@login_required
def delete_company(company_id):
    company = Company.query.get_or_404(company_id)
    
    # Check if company has subcompanies
    if company.subcompanies:
        flash('Cannot delete company with existing subcompanies. Please delete subcompanies first.', 'error')
        return redirect(url_for('clients.company_detail', company_id=company.id))
    
    # Check if company is a subcompany of another company
    if company.parent_company_id:
        flash('Cannot delete a subcompany from this page. Use the companies list instead.', 'error')
        return redirect(url_for('clients.company_detail', company_id=company.id))
    
    try:
        # Delete the company - related entities will cascade delete
        db.session.delete(company)
        db.session.commit()
        flash(f'Company "{company.name}" and all related data have been deleted successfully!', 'success')
        return redirect(url_for('clients.companies'))
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while deleting the company. Please try again.', 'error')
        return redirect(url_for('clients.company_detail', company_id=company.id))

@bp.route('/company/<int:company_id>/agreement', methods=['GET', 'POST'])
@login_required
def upload_agreement(company_id):
    company = Company.query.get_or_404(company_id)
    form = AgreementForm()
    
    if form.validate_on_submit():
        if form.file.data:
            filename = secure_filename(form.file.data.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{company_id}_{timestamp}_{filename}"
            file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            form.file.data.save(file_path)
            
            agreement = Agreement(
                company_id=company_id,
                type=form.type.data,
                filename=form.file.data.filename,
                file_path=filename,
                valid_until=form.valid_until.data,
                uploaded_by_id=current_user.id
            )
            db.session.add(agreement)
            db.session.commit()
            flash('Agreement uploaded successfully!', 'success')
            return redirect(url_for('clients.company_detail', company_id=company_id))
    
    return render_template('clients/upload_agreement.html', form=form, company=company)

@bp.route('/brands')
@login_required
def brands():
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    return render_template('clients/brands.html', brands=brands)

@bp.route('/brand/new', methods=['GET', 'POST'])
@login_required
def new_brand():
    form = BrandForm()
    
    # Pre-select company if company_id is in query params
    company_id = request.args.get('company_id', type=int)
    if company_id and request.method == 'GET':
        form.company_id.data = company_id
    
    if form.validate_on_submit():
        brand = Brand(
            name=form.name.data,
            company_id=form.company_id.data,
            status=form.status.data
        )
        db.session.add(brand)
        db.session.commit()
        flash('Brand created successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand.id))
    return render_template('clients/brand_form.html', form=form, title='New Brand')

@bp.route('/brand/<int:brand_id>')
@login_required
def brand_detail(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    return render_template('clients/brand_detail.html', brand=brand)

@bp.route('/brand/<int:brand_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_brand(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = BrandForm(obj=brand)
    
    if form.validate_on_submit():
        brand.name = form.name.data
        brand.company_id = form.company_id.data
        brand.status = form.status.data
        db.session.commit()
        flash('Brand updated successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand.id))
    
    return render_template('clients/brand_form.html', form=form, title='Edit Brand', brand=brand)

@bp.route('/brand/<int:brand_id>/delete', methods=['POST'])
@login_required
def delete_brand(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    company_id = brand.company_id
    
    try:
        # Delete all related records first (subbrands will be cascade deleted)
        db.session.delete(brand)
        db.session.commit()
        flash('Brand deleted successfully!', 'success')
        return redirect(url_for('clients.company_detail', company_id=company_id))
    except Exception as e:
        db.session.rollback()
        # Log the actual error for debugging
        print(f"Error deleting brand {brand_id}: {str(e)}")
        flash(f'Error deleting brand: {str(e)}', 'error')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))

@bp.route('/brand/<int:brand_id>/subbrand/new', methods=['GET', 'POST'])
@login_required
def new_subbrand(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = SubbrandForm()
    
    if form.validate_on_submit():
        subbrand = Subbrand(
            name=form.name.data,
            brand_id=brand_id
        )
        db.session.add(subbrand)
        db.session.commit()
        flash('Subbrand added successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    return render_template('clients/subbrand_form.html', form=form, brand=brand)

@bp.route('/brand/<int:brand_id>/assign-contact', methods=['GET', 'POST'])
@login_required
def assign_contact(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'existing':
            # Assign existing contacts
            contact_ids = request.form.getlist('contact_ids')
            for contact_id in contact_ids:
                contact = ClientContact.query.get(contact_id)
                if contact and contact not in brand.contacts:
                    brand.contacts.append(contact)
            
            db.session.commit()
            flash('Contacts assigned successfully!', 'success')
            return redirect(url_for('clients.brand_detail', brand_id=brand_id))
        
        elif action == 'new':
            # Redirect to new contact form with brand pre-selected
            return redirect(url_for('clients.new_contact', brand_id=brand_id))
    
    # Get all contacts not already assigned to this brand
    assigned_contact_ids = [c.id for c in brand.contacts]
    available_contacts = ClientContact.query.filter(
        ~ClientContact.id.in_(assigned_contact_ids) if assigned_contact_ids else True
    ).order_by(ClientContact.last_name, ClientContact.first_name).all()
    
    return render_template('clients/assign_contact.html', brand=brand, available_contacts=available_contacts)

@bp.route('/brand/<int:brand_id>/team', methods=['GET', 'POST'])
@login_required
def assign_team(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = BrandTeamForm()
    
    if form.validate_on_submit():
        BrandTeam.query.filter_by(brand_id=brand_id).delete()
        
        key_responsible_id = int(form.key_responsible_id.data) if form.key_responsible_id.data and form.key_responsible_id.data != '0' else None
        
        for user_id in form.team_members.data:
            assignment = BrandTeam(
                brand_id=brand_id,
                team_member_id=user_id,
                is_key_responsible=(user_id == key_responsible_id)
            )
            db.session.add(assignment)
        
        db.session.commit()
        flash('Team assigned successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    elif request.method == 'GET':
        current_assignments = BrandTeam.query.filter_by(brand_id=brand_id).all()
        form.team_members.data = [a.team_member_id for a in current_assignments]
        key_responsible = next((a for a in current_assignments if a.is_key_responsible), None)
        if key_responsible:
            form.key_responsible_id.data = str(key_responsible.team_member_id)
    
    return render_template('clients/assign_team.html', form=form, brand=brand)

@bp.route('/brand/<int:brand_id>/planning', methods=['GET', 'POST'])
@login_required
def planning_info(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = PlanningInfoForm()
    
    if form.validate_on_submit():
        planning = PlanningInfo(
            brand_id=brand_id,
            comments=form.comments.data,
            kpis='',  # Keep empty for backward compatibility
            created_by_id=current_user.id
        )
        db.session.add(planning)
        db.session.flush()  # Get planning ID before handling attachments
        
        # Handle file attachments
        if form.attachments.data:
            for file in form.attachments.data:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"planning_{planning.id}_{timestamp}_{filename}"
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    
                    attachment = PlanningAttachment(
                        planning_info_id=planning.id,
                        filename=file.filename,
                        file_path=filename
                    )
                    db.session.add(attachment)
        
        db.session.commit()
        flash('Planning information added!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    planning_records = brand.planning_info
    return render_template('clients/planning_info.html', form=form, brand=brand, planning_records=planning_records)

@bp.route('/brand/<int:brand_id>/status', methods=['GET', 'POST'])
@login_required
def add_status_update(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = StatusUpdateForm()
    
    if form.validate_on_submit():
        update = StatusUpdate(
            brand_id=brand_id,
            date=form.date.data,
            comment=form.comment.data,
            evaluation=form.evaluation.data,
            created_by_id=current_user.id
        )
        db.session.add(update)
        db.session.commit()
        flash('Status update added!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    return render_template('clients/status_update.html', form=form, brand=brand)

@bp.route('/contacts')
@login_required
def contacts():
    # Get filter parameters
    brand_id = request.args.get('brand_id', type=int)
    company_id = request.args.get('company_id', type=int)
    contact_type = request.args.get('contact_type', '').strip()
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = ClientContact.query
    
    # Apply filters
    if brand_id:
        query = query.join(ClientContact.brands).filter(Brand.id == brand_id)
    if company_id:
        query = query.join(ClientContact.brands).filter(Brand.company_id == company_id)
    if contact_type:
        query = query.filter(ClientContact.contact_type == contact_type)
    if search:
        search_filter = f'%{search}%'
        query = query.filter(db.or_(
            ClientContact.first_name.ilike(search_filter),
            ClientContact.last_name.ilike(search_filter),
            ClientContact.email.ilike(search_filter),
            ClientContact.phone.ilike(search_filter)
        ))
    
    # Get paginated contacts
    pagination = query.order_by(ClientContact.last_name, ClientContact.first_name).paginate(
        page=page, per_page=per_page, error_out=False)
    contacts = pagination.items
    
    # Get all brands and companies for filter dropdowns
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    companies = Company.query.order_by(Company.name).all()
    
    return render_template('clients/contacts.html', 
                         contacts=contacts,
                         brands=brands,
                         companies=companies,
                         pagination=pagination,
                         selected_brand_id=brand_id,
                         selected_company_id=company_id,
                         selected_contact_type=contact_type,
                         search_query=search)

@bp.route('/contact/new', methods=['GET', 'POST'])
@bp.route('/brand/<int:brand_id>/contact/new', methods=['GET', 'POST'])
@login_required
def new_contact(brand_id=None):
    form = ClientContactForm()
    
    # Initialize brands data to empty list if None
    if form.brands.data is None:
        form.brands.data = []
    
    # If coming from a brand page, pre-select that brand
    if brand_id and request.method == 'GET':
        brand = Brand.query.get_or_404(brand_id)
        form.brands.data = [brand_id]
    
    if form.validate_on_submit():
        contact = ClientContact(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            phone=form.phone.data,
            linkedin_url=form.linkedin_url.data,
            birthday_month=form.birthday_month.data if form.birthday_month.data != 0 else None,
            birthday_day=form.birthday_day.data,
            responsibility_description=form.responsibility_description.data,
            should_get_gift=form.should_get_gift.data,
            receive_newsletter=form.receive_newsletter.data,
            status=form.status.data,
            contact_type=form.contact_type.data
        )
        
        for brand_id in form.brands.data:
            brand = Brand.query.get(brand_id)
            if brand:
                contact.brands.append(brand)
        
        db.session.add(contact)
        db.session.commit()
        flash('Contact created successfully!', 'success')
        
        # If created from brand page, redirect back to brand
        if brand_id:
            return redirect(url_for('clients.brand_detail', brand_id=brand_id))
        return redirect(url_for('clients.contact_detail', contact_id=contact.id))
    
    return render_template('clients/contact_form.html', form=form, title='New Contact', brand_id=brand_id)

@bp.route('/contact/<int:contact_id>')
@login_required
def contact_detail(contact_id):
    contact = ClientContact.query.get_or_404(contact_id)
    return render_template('clients/contact_detail.html', contact=contact)

@bp.route('/contact/<int:contact_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_contact(contact_id):
    contact = ClientContact.query.get_or_404(contact_id)
    form = ClientContactForm(contact=contact)
    
    if form.validate_on_submit():
        contact.first_name = form.first_name.data
        contact.last_name = form.last_name.data
        contact.email = form.email.data
        contact.phone = form.phone.data
        contact.linkedin_url = form.linkedin_url.data
        contact.birthday_month = form.birthday_month.data if form.birthday_month.data != 0 else None
        contact.birthday_day = form.birthday_day.data
        contact.responsibility_description = form.responsibility_description.data
        contact.should_get_gift = form.should_get_gift.data
        contact.receive_newsletter = form.receive_newsletter.data
        contact.status = form.status.data
        contact.contact_type = form.contact_type.data
        
        contact.brands.clear()
        for brand_id in form.brands.data:
            brand = Brand.query.get(brand_id)
            if brand:
                contact.brands.append(brand)
        
        db.session.commit()
        flash('Contact updated successfully!', 'success')
        return redirect(url_for('clients.contact_detail', contact_id=contact.id))
    
    elif request.method == 'GET':
        form.first_name.data = contact.first_name
        form.last_name.data = contact.last_name
        form.email.data = contact.email
        form.phone.data = contact.phone
        form.linkedin_url.data = contact.linkedin_url
        form.birthday_month.data = contact.birthday_month or 0
        form.birthday_day.data = contact.birthday_day
        form.responsibility_description.data = contact.responsibility_description
        form.should_get_gift.data = contact.should_get_gift
        form.receive_newsletter.data = contact.receive_newsletter
        form.status.data = contact.status
        form.contact_type.data = contact.contact_type
        form.brands.data = [b.id for b in contact.brands]
    
    return render_template('clients/contact_form.html', form=form, title='Edit Contact', contact=contact)

@bp.route('/company/<int:company_id>/commitment', methods=['GET', 'POST'])
@login_required
def add_commitment(company_id):
    company = Company.query.get_or_404(company_id)
    form = CommitmentForm()
    
    if form.validate_on_submit():
        commitment = Commitment(
            company_id=company_id,
            media_group_id=form.media_group_id.data,
            year=form.year.data,
            amount=form.amount.data,
            currency=form.currency.data
        )
        db.session.add(commitment)
        try:
            db.session.commit()
            flash('Commitment added successfully!', 'success')
        except:
            db.session.rollback()
            flash('A commitment for this media group and year already exists!', 'error')
        return redirect(url_for('clients.company_detail', company_id=company_id))
    
    return render_template('clients/commitment_form.html', form=form, company=company)

@bp.route('/media-groups')
@login_required
def media_groups():
    media_groups = MediaGroup.query.order_by(MediaGroup.name).all()
    return render_template('clients/media_groups.html', media_groups=media_groups)

@bp.route('/media-group/new', methods=['GET', 'POST'])
@login_required
def new_media_group():
    form = MediaGroupForm()
    if form.validate_on_submit():
        media_group = MediaGroup(name=form.name.data)
        db.session.add(media_group)
        try:
            db.session.commit()
            flash('Media group created successfully!', 'success')
            return redirect(url_for('clients.media_groups'))
        except:
            db.session.rollback()
            flash('A media group with this name already exists!', 'error')
    
    return render_template('clients/media_group_form.html', form=form, title='New Media Group')

@bp.route('/brand/<int:brand_id>/meeting', methods=['GET', 'POST'])
@login_required
def add_meeting(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = KeyMeetingForm()
    
    if form.validate_on_submit():
        meeting = KeyMeeting(
            brand_id=brand_id,
            date=form.date.data,
            comment=form.comment.data,
            created_by_id=current_user.id
        )
        db.session.add(meeting)
        db.session.flush()
        
        # Handle file attachments
        if form.attachments.data:
            for file in form.attachments.data:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"meeting_{meeting.id}_{timestamp}_{filename}"
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    
                    attachment = MeetingAttachment(
                        meeting_id=meeting.id,
                        filename=file.filename,
                        file_path=filename
                    )
                    db.session.add(attachment)
        
        db.session.commit()
        flash('Meeting added successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    return render_template('clients/meeting_form.html', form=form, brand=brand)

@bp.route('/brand/<int:brand_id>/link', methods=['GET', 'POST'])
@login_required
def add_link(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = KeyLinkForm()
    
    if form.validate_on_submit():
        link = KeyLink(
            brand_id=brand_id,
            url=form.url.data,
            comment=form.comment.data,
            created_by_id=current_user.id
        )
        db.session.add(link)
        db.session.commit()
        flash('Link added successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    return render_template('clients/link_form.html', form=form, brand=brand)

@bp.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@bp.route('/birthdays')
@login_required
def birthdays():
    # Get current month and next 2 months
    today = datetime.now()
    current_month = today.month
    
    # Get contacts with birthdays in the next 3 months
    upcoming_contacts = []
    for i in range(3):
        month = ((current_month - 1 + i) % 12) + 1
        contacts = ClientContact.query.filter(
            ClientContact.birthday_month == month,
            ClientContact.should_get_gift == True,
            ClientContact.status == 'active'
        ).order_by(ClientContact.birthday_day).all()
        
        for contact in contacts:
            # Check if gift already logged this year
            current_year_gift = Gift.query.filter_by(
                contact_id=contact.id, 
                year=today.year
            ).first()
            
            contact_info = {
                'contact': contact,
                'month': month,
                'current_year_gift': current_year_gift
            }
            upcoming_contacts.append(contact_info)
    
    # Sort by month and day
    upcoming_contacts.sort(key=lambda x: (x['month'], x['contact'].birthday_day or 0))
    
    return render_template('clients/birthdays.html', upcoming_contacts=upcoming_contacts, current_year=today.year)

@bp.route('/contact/<int:contact_id>/gift', methods=['GET', 'POST'])
@login_required
def add_gift(contact_id):
    contact = ClientContact.query.get_or_404(contact_id)
    form = GiftForm()
    
    # Pre-fill current year
    if request.method == 'GET':
        form.year.data = datetime.now().year
    
    if form.validate_on_submit():
        # Check if gift already exists for this year
        existing_gift = Gift.query.filter_by(
            contact_id=contact_id,
            year=form.year.data
        ).first()
        
        if existing_gift:
            flash('A gift has already been recorded for this contact in this year!', 'error')
        else:
            gift = Gift(
                contact_id=contact_id,
                year=form.year.data,
                gift_description=form.gift_description.data,
                gift_value=form.gift_value.data,
                sent_date=form.sent_date.data,
                notes=form.notes.data,
                created_by_id=current_user.id
            )
            db.session.add(gift)
            db.session.commit()
            flash('Gift recorded successfully!', 'success')
            return redirect(url_for('clients.birthdays'))
    
    # Get gift history for this contact
    gifts = Gift.query.filter_by(contact_id=contact_id).order_by(Gift.year.desc()).all()
    
    return render_template('clients/gift_form.html', form=form, contact=contact, gifts=gifts)

@bp.route('/status-updates')
@login_required
def status_updates():
    # Get filter parameters
    brand_id = request.args.get('brand_id', type=int)
    evaluation = request.args.get('evaluation')
    created_by_id = request.args.get('created_by_id', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = StatusUpdate.query.join(Brand).join(Company)
    
    if brand_id:
        query = query.filter(StatusUpdate.brand_id == brand_id)
    if evaluation:
        query = query.filter(StatusUpdate.evaluation == evaluation)
    if created_by_id:
        query = query.filter(StatusUpdate.created_by_id == created_by_id)
    
    # Get paginated status updates ordered by date
    pagination = query.order_by(StatusUpdate.date.desc()).paginate(
        page=page, per_page=per_page, error_out=False)
    updates = pagination.items
    
    # Get all brands for filter dropdown
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    
    # Get all users who have created status updates for filter dropdown
    creators = db.session.query(User).join(StatusUpdate, User.id == StatusUpdate.created_by_id).distinct().order_by(User.first_name, User.last_name).all()
    
    return render_template('clients/status_updates.html', 
                         updates=updates, 
                         brands=brands,
                         creators=creators,
                         pagination=pagination,
                         selected_brand_id=brand_id,
                         selected_evaluation=evaluation,
                         selected_created_by_id=created_by_id)

@bp.route('/status-update/new', methods=['GET', 'POST'])
@login_required
def new_status_update():
    # Create a custom form class with brand_id field
    class StatusUpdateFormWithBrand(StatusUpdateForm):
        brand_id = SelectField('Brand', coerce=int, validators=[DataRequired()])
    
    form = StatusUpdateFormWithBrand()
    
    # Add brand choices to form
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    form.brand_id.choices = [(b.id, f"{b.name} ({b.company.name})") for b in brands]
    
    if form.validate_on_submit():
        update = StatusUpdate(
            brand_id=form.brand_id.data,
            date=form.date.data,
            comment=form.comment.data,
            evaluation=form.evaluation.data,
            created_by_id=current_user.id
        )
        db.session.add(update)
        db.session.commit()
        flash('Status update added successfully!', 'success')
        return redirect(url_for('clients.status_updates'))
    
    # Set default date to today
    if request.method == 'GET':
        form.date.data = datetime.now().date()
    
    return render_template('clients/status_update_form.html', form=form, title='New Status Update')

@bp.route('/tasks')
@login_required
def tasks():
    # Get all active brand tasks with upcoming due dates
    today = datetime.now().date()
    
    # Get all active tasks
    active_tasks = BrandTask.query.filter_by(is_active=True).join(Brand).join(TaskTemplate).all()
    
    # Calculate due dates and organize by brand
    tasks_by_brand = {}
    for task in active_tasks:
        next_due = task.get_next_due_date()
        
        # Only show tasks that are due within the next 90 days
        if next_due and (next_due - today).days <= 90:
            if task.brand_id not in tasks_by_brand:
                tasks_by_brand[task.brand_id] = {
                    'brand': task.brand,
                    'tasks': []
                }
            
            # Check if already completed for this period
            completed = TaskCompletion.query.filter(
                TaskCompletion.brand_task_id == task.id,
                TaskCompletion.completion_date >= next_due - timedelta(days=7)
            ).first()
            
            tasks_by_brand[task.brand_id]['tasks'].append({
                'task': task,
                'next_due': next_due,
                'is_overdue': next_due < today,
                'is_completed': completed is not None,
                'completion': completed
            })
    
    # Sort tasks by due date within each brand
    for brand_id in tasks_by_brand:
        tasks_by_brand[brand_id]['tasks'].sort(key=lambda x: x['next_due'])
    
    return render_template('clients/tasks.html', tasks_by_brand=tasks_by_brand)

@bp.route('/brand/<int:brand_id>/tasks')
@login_required
def brand_tasks(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    # Get all task templates
    templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
    
    # Get brand's current tasks
    brand_tasks = BrandTask.query.filter_by(brand_id=brand_id).all()
    assigned_template_ids = [bt.task_template_id for bt in brand_tasks]
    
    # Get available templates (not yet assigned)
    available_templates = [t for t in templates if t.id not in assigned_template_ids]
    
    # Calculate next due dates for active tasks
    tasks_with_due_dates = []
    for bt in brand_tasks:
        if bt.is_active:
            next_due = bt.get_next_due_date()
            last_completion = TaskCompletion.query.filter_by(
                brand_task_id=bt.id
            ).order_by(TaskCompletion.completion_date.desc()).first()
            
            tasks_with_due_dates.append({
                'task': bt,
                'next_due': next_due,
                'last_completion': last_completion,
                'is_overdue': next_due < datetime.now().date() if next_due else False
            })
    
    # Sort by next due date
    tasks_with_due_dates.sort(key=lambda x: x['next_due'] if x['next_due'] else datetime.max.date())
    
    return render_template('clients/brand_tasks.html', 
                         brand=brand, 
                         tasks=tasks_with_due_dates,
                         available_templates=available_templates)

@bp.route('/brand/<int:brand_id>/task/new', methods=['GET', 'POST'])
@login_required
def new_brand_task(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = BrandTaskForm()
    
    # Get available task templates
    existing_tasks = BrandTask.query.filter_by(brand_id=brand_id).all()
    assigned_template_ids = [bt.task_template_id for bt in existing_tasks]
    
    available_templates = TaskTemplate.query.filter(
        ~TaskTemplate.id.in_(assigned_template_ids)
    ).order_by(TaskTemplate.name).all()
    
    form.task_template_id.choices = [(t.id, t.name) for t in available_templates]
    
    if form.validate_on_submit():
        task = BrandTask(
            brand_id=brand_id,
            task_template_id=form.task_template_id.data,
            frequency=form.frequency.data,
            start_date=form.start_date.data,
            created_by_id=current_user.id
        )
        db.session.add(task)
        db.session.commit()
        flash('Task added successfully!', 'success')
        return redirect(url_for('clients.brand_tasks', brand_id=brand_id))
    
    # Set default start date to today
    if request.method == 'GET':
        form.start_date.data = datetime.now().date()
    
    return render_template('clients/brand_task_form.html', form=form, brand=brand)

@bp.route('/brand-task/<int:task_id>/complete', methods=['GET', 'POST'])
@login_required
def complete_task(task_id):
    task = BrandTask.query.get_or_404(task_id)
    form = TaskCompletionForm()
    
    if form.validate_on_submit():
        completion = TaskCompletion(
            brand_task_id=task_id,
            completion_date=form.completion_date.data,
            notes=form.notes.data,
            completed_by_id=current_user.id
        )
        db.session.add(completion)
        db.session.commit()
        flash('Task marked as complete!', 'success')
        return redirect(url_for('clients.brand_tasks', brand_id=task.brand_id))
    
    # Set default date to today
    if request.method == 'GET':
        form.completion_date.data = datetime.now().date()
    
    return render_template('clients/task_completion_form.html', form=form, task=task)

@bp.route('/brand-task/<int:task_id>/toggle-active', methods=['POST'])
@login_required
def toggle_task_active(task_id):
    task = BrandTask.query.get_or_404(task_id)
    task.is_active = not task.is_active
    db.session.commit()
    
    status = 'activated' if task.is_active else 'deactivated'
    flash(f'Task {status} successfully!', 'success')
    return redirect(url_for('clients.brand_tasks', brand_id=task.brand_id))

@bp.route('/task-templates')
@login_required
def task_templates():
    templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
    return render_template('clients/task_templates.html', templates=templates)

@bp.route('/task-template/new', methods=['GET', 'POST'])
@login_required
def new_task_template():
    form = TaskTemplateForm()
    
    if form.validate_on_submit():
        template = TaskTemplate(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(template)
        try:
            db.session.commit()
            flash('Task template created successfully!', 'success')
            return redirect(url_for('clients.task_templates'))
        except:
            db.session.rollback()
            flash('A task template with this name already exists!', 'error')
    
    return render_template('clients/task_template_form.html', form=form, title='New Task Template')

@bp.route('/company/<int:company_id>/subcompany/new', methods=['GET', 'POST'])
@login_required
def new_subcompany(company_id):
    parent_company = Company.query.get_or_404(company_id)
    form = SubcompanyForm()
    
    if form.validate_on_submit():
        subcompany = Company(
            name=form.name.data,
            vat_code=form.vat_code.data,
            registration_number=form.registration_number.data,
            address=form.address.data,
            bank_account=form.bank_account.data,
            parent_company_id=company_id,
            status='active'
        )
        db.session.add(subcompany)
        db.session.commit()
        flash('Subcompany created successfully!', 'success')
        return redirect(url_for('clients.company_detail', company_id=company_id))
    
    return render_template('clients/subcompany_form.html', form=form, parent_company=parent_company)

@bp.route('/company/<int:company_id>/subcompany/<int:subcompany_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_subcompany(company_id, subcompany_id):
    parent_company = Company.query.get_or_404(company_id)
    subcompany = Company.query.get_or_404(subcompany_id)
    
    if subcompany.parent_company_id != company_id:
        abort(404)
    
    form = SubcompanyForm()
    
    if form.validate_on_submit():
        subcompany.name = form.name.data
        subcompany.vat_code = form.vat_code.data
        subcompany.registration_number = form.registration_number.data
        subcompany.address = form.address.data
        subcompany.bank_account = form.bank_account.data
        subcompany.updated_at = datetime.utcnow()
        db.session.commit()
        flash('Subcompany updated successfully!', 'success')
        return redirect(url_for('clients.company_detail', company_id=company_id))
    
    elif request.method == 'GET':
        form.name.data = subcompany.name
        form.vat_code.data = subcompany.vat_code
        form.registration_number.data = subcompany.registration_number
        form.address.data = subcompany.address
        form.bank_account.data = subcompany.bank_account
    
    return render_template('clients/subcompany_form.html', form=form, parent_company=parent_company, subcompany=subcompany)

@bp.route('/subcompany/<int:subcompany_id>/delete', methods=['POST'])
@login_required
def delete_subcompany(subcompany_id):
    subcompany = Company.query.get_or_404(subcompany_id)
    
    # Verify this is actually a subcompany
    if not subcompany.parent_company_id:
        flash('This is not a subcompany.', 'error')
        return redirect(url_for('clients.companies'))
    
    parent_company_id = subcompany.parent_company_id
    
    # Check if subcompany has its own subcompanies
    if subcompany.subcompanies:
        flash('Cannot delete subcompany with existing sub-subcompanies.', 'error')
        return redirect(url_for('clients.companies'))
    
    try:
        # Delete the subcompany - related entities will cascade delete
        db.session.delete(subcompany)
        db.session.commit()
        flash(f'Subcompany "{subcompany.name}" and all related data have been deleted successfully!', 'success')
        return redirect(url_for('clients.companies'))
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while deleting the subcompany. Please try again.', 'error')
        return redirect(url_for('clients.companies'))

@bp.route('/invoices')
@login_required
def invoices():
    # Get filter and sort parameters
    brand_id = request.args.get('brand_id', type=int)
    company_id = request.args.get('company_id', type=int)
    sort_by = request.args.get('sort_by', 'date')  # date or amount
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = Invoice.query.join(Brand).join(Company)
    
    # Apply filters
    if brand_id:
        query = query.filter(Invoice.brand_id == brand_id)
    if company_id:
        query = query.filter(Brand.company_id == company_id)
    
    # Apply sorting
    if sort_by == 'amount':
        query = query.order_by(Invoice.total_amount.desc())
    else:  # default to date
        query = query.order_by(Invoice.invoice_date.desc())
    
    # Get paginated invoices
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    invoices = pagination.items
    
    # Get all brands and companies for filter dropdowns
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    companies = Company.query.order_by(Company.name).all()
    
    # Calculate total amount for current page
    page_total = sum(invoice.total_amount for invoice in invoices)
    
    return render_template('clients/invoices.html', 
                         invoices=invoices,
                         brands=brands,
                         companies=companies,
                         pagination=pagination,
                         selected_brand_id=brand_id,
                         selected_company_id=company_id,
                         sort_by=sort_by,
                         page_total=page_total)

@bp.route('/brand/<int:brand_id>/invoice/new', methods=['GET', 'POST'])
@login_required
def new_invoice(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = InvoiceForm()
    
    # Get all companies (main company + subcompanies)
    companies = [brand.company]
    if brand.company.subcompanies:
        companies.extend(brand.company.subcompanies)
    form.company_id.choices = [(c.id, c.name) for c in companies]
    
    if form.validate_on_submit():
        invoice = Invoice(
            brand_id=brand_id,
            company_id=form.company_id.data,
            invoice_date=form.invoice_date.data,
            short_info=form.short_info.data,
            total_amount=form.total_amount.data,
            created_by_id=current_user.id
        )
        
        db.session.add(invoice)
        db.session.flush()  # Get the invoice ID before saving files
        
        # Handle multiple file uploads
        if form.files.data:
            file_count = 0
            for file in form.files.data:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"invoice_{invoice.id}_{timestamp}_{filename}"
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    
                    attachment = InvoiceAttachment(
                        invoice_id=invoice.id,
                        filename=file.filename,
                        file_path=filename
                    )
                    db.session.add(attachment)
                    file_count += 1
                    
                    # For backward compatibility, store first file in invoice table
                    if file_count == 1:
                        invoice.filename = file.filename
                        invoice.file_path = filename
        
        db.session.commit()
        flash('Invoice registered successfully!', 'success')
        return redirect(url_for('clients.brand_detail', brand_id=brand_id))
    
    return render_template('clients/invoice_form.html', form=form, brand=brand)

@bp.route('/invoice/<int:invoice_id>/download')
@login_required
def download_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    if invoice.file_path:
        return send_from_directory(current_app.config['UPLOAD_FOLDER'], invoice.file_path, 
                                 as_attachment=True, download_name=invoice.filename)

@bp.route('/brands/export')
@login_required
def export_brands():
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Brands"
    
    # Add headers
    headers = ['Company', 'Brand', 'Subbrands', 'Status', 'Key Responsible', 'Last Update', 'Risk Level']
    ws.append(headers)
    
    # Get all brands
    brands = Brand.query.join(Company).order_by(Company.name, Brand.name).all()
    
    # Add data rows
    for brand in brands:
        # Get key responsible
        key_responsible = None
        for tm in brand.team_members:
            if tm.is_key_responsible:
                key_responsible = f"{tm.team_member.first_name} {tm.team_member.last_name}"
                break
        
        # Get latest status update
        latest_update = None
        risk_level = None
        if brand.status_updates:
            latest_update = max(brand.status_updates, key=lambda x: x.date)
            risk_level = latest_update.evaluation
            latest_update = latest_update.date.strftime('%Y-%m-%d')
        
        # Get subbrands
        subbrands = ', '.join([sb.name for sb in brand.subbrands])
        
        row = [
            brand.company.name,
            brand.name,
            subbrands,
            brand.status,
            key_responsible or 'Not assigned',
            latest_update or 'Never updated',
            risk_level or 'No evaluation'
        ]
        ws.append(row)
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=brands_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@bp.route('/contacts/export')
@login_required
def export_contacts():
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Add headers
    headers = ['First Name', 'Last Name', 'Email', 'Phone', 'LinkedIn', 'Birthday', 'Brands', 'Should Get Gift', 'Newsletter', 'Status']
    ws.append(headers)
    
    # Get all contacts
    contacts = ClientContact.query.order_by(ClientContact.last_name, ClientContact.first_name).all()
    
    # Add data rows
    for contact in contacts:
        # Get associated brands
        brands = ', '.join([f"{b.name} ({b.company.name})" for b in contact.brands])
        
        # Format birthday
        birthday = ''
        if contact.birthday_month and contact.birthday_day:
            birthday = f"{contact.birthday_month:02d}-{contact.birthday_day:02d}"
        
        row = [
            contact.first_name,
            contact.last_name,
            contact.email,
            contact.phone or '',
            contact.linkedin_url or '',
            birthday,
            brands,
            'Yes' if contact.should_get_gift else 'No',
            'Yes' if contact.receive_newsletter else 'No',
            contact.status
        ]
        ws.append(row)
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=contacts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@bp.route('/companies/export')
@login_required
def export_companies():
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    
    # Add headers
    headers = ['Company Name', 'VAT Code', 'Registration Number', 'Address', 'Bank Account', 
               'Agency Fees', 'Parent Company', 'Brands', 'Active Service Agreement', 
               'Active Data Agreement', 'Status']
    ws.append(headers)
    
    # Get all companies
    companies = Company.query.order_by(Company.name).all()
    
    # Add data rows
    for company in companies:
        # Get brands
        brands = ', '.join([b.name for b in company.brands])
        
        # Check agreements
        has_service = 'No'
        has_data = 'No'
        for agreement in company.agreements:
            if agreement.type == 'service' and (not agreement.valid_until or agreement.valid_until >= datetime.now().date()):
                has_service = 'Yes'
            elif agreement.type == 'data' and (not agreement.valid_until or agreement.valid_until >= datetime.now().date()):
                has_data = 'Yes'
        
        row = [
            company.name,
            company.vat_code or '',
            company.registration_number or '',
            company.address or '',
            company.bank_account or '',
            company.agency_fees or '',
            company.parent_company.name if company.parent_company else '',
            brands,
            has_service,
            has_data,
            company.status
        ]
        ws.append(row)
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=companies_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@bp.route('/brand/<int:brand_id>/media-planning')
@login_required
def media_planning(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    # Get filter parameters
    year = request.args.get('year', type=int, default=datetime.now().year)
    quarter = request.args.get('quarter', type=int)
    
    # Build query
    query = MediaPlan.query.filter_by(brand_id=brand_id)
    
    if year:
        query = query.filter_by(year=year)
    if quarter:
        query = query.filter_by(quarter=quarter)
    
    media_plans = query.order_by(MediaPlan.year.desc(), MediaPlan.quarter.desc(), MediaPlan.media_type).all()
    
    # Get available years for filter
    years = db.session.query(MediaPlan.year).filter_by(brand_id=brand_id).distinct().order_by(MediaPlan.year.desc()).all()
    years = [y[0] for y in years]
    if not years:
        years = [datetime.now().year]
    
    # Calculate totals
    total_planned = sum(mp.planned_budget or 0 for mp in media_plans)
    total_actual = sum(mp.actual_spend or 0 for mp in media_plans)
    
    return render_template('clients/media_planning.html', 
                         brand=brand,
                         media_plans=media_plans,
                         selected_year=year,
                         selected_quarter=quarter,
                         available_years=years,
                         total_planned=total_planned,
                         total_actual=total_actual)

@bp.route('/brand/<int:brand_id>/media-planning/add', methods=['GET', 'POST'])
@login_required
def add_media_plan(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    form = MediaPlanForm()
    
    # Set year choices
    current_year = datetime.now().year
    form.year.choices = [(y, str(y)) for y in range(current_year - 1, current_year + 3)]
    
    if form.validate_on_submit():
        media_plan = MediaPlan(
            brand_id=brand_id,
            year=form.year.data,
            quarter=form.quarter.data,
            media_type=form.media_type.data,
            channel_name=form.channel_name.data,
            planned_budget=form.planned_budget.data,
            actual_spend=form.actual_spend.data,
            notes=form.notes.data
        )
        db.session.add(media_plan)
        db.session.commit()
        flash('Media plan added successfully!', 'success')
        return redirect(url_for('clients.media_planning', brand_id=brand_id))
    
    # Set default year
    if request.method == 'GET':
        form.year.data = current_year
    
    return render_template('clients/media_plan_form.html', form=form, brand=brand, title='Add Media Plan')

@bp.route('/brand/<int:brand_id>/media-planning/<int:plan_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_media_plan(brand_id, plan_id):
    brand = Brand.query.get_or_404(brand_id)
    media_plan = MediaPlan.query.get_or_404(plan_id)
    
    if media_plan.brand_id != brand_id:
        abort(404)
    
    form = MediaPlanForm()
    
    # Set year choices
    current_year = datetime.now().year
    form.year.choices = [(y, str(y)) for y in range(current_year - 2, current_year + 3)]
    
    if form.validate_on_submit():
        media_plan.year = form.year.data
        media_plan.quarter = form.quarter.data
        media_plan.media_type = form.media_type.data
        media_plan.channel_name = form.channel_name.data
        media_plan.planned_budget = form.planned_budget.data
        media_plan.actual_spend = form.actual_spend.data
        media_plan.notes = form.notes.data
        
        db.session.commit()
        flash('Media plan updated successfully!', 'success')
        return redirect(url_for('clients.media_planning', brand_id=brand_id))
    
    elif request.method == 'GET':
        form.year.data = media_plan.year
        form.quarter.data = media_plan.quarter
        form.media_type.data = media_plan.media_type
        form.channel_name.data = media_plan.channel_name
        form.planned_budget.data = media_plan.planned_budget
        form.actual_spend.data = media_plan.actual_spend
        form.notes.data = media_plan.notes
    
    return render_template('clients/media_plan_form.html', form=form, brand=brand, title='Edit Media Plan')

@bp.route('/brand/<int:brand_id>/media-planning/<int:plan_id>/delete', methods=['POST'])
@login_required
def delete_media_plan(brand_id, plan_id):
    media_plan = MediaPlan.query.get_or_404(plan_id)
    
    if media_plan.brand_id != brand_id:
        abort(404)
    
    db.session.delete(media_plan)
    db.session.commit()
    flash('Media plan deleted successfully!', 'success')
    return redirect(url_for('clients.media_planning', brand_id=brand_id))

@bp.route('/brand/<int:brand_id>/digital-info')
@login_required
def digital_info(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    # Get or create digital info for this brand
    info = DigitalInfo.query.filter_by(brand_id=brand_id).first()
    if not info:
        info = DigitalInfo(brand_id=brand_id)
        db.session.add(info)
        db.session.commit()
    
    return render_template('clients/digital_info.html', brand=brand, info=info)

@bp.route('/brand/<int:brand_id>/digital-info/edit', methods=['GET', 'POST'])
@login_required
def edit_digital_info(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    # Get or create digital info
    info = DigitalInfo.query.filter_by(brand_id=brand_id).first()
    if not info:
        info = DigitalInfo(brand_id=brand_id)
        db.session.add(info)
        db.session.commit()
    
    form = DigitalInfoForm()
    
    if form.validate_on_submit():
        info.digital_planning_info = form.digital_planning_info.data
        info.digital_adops_info = form.digital_adops_info.data
        info.digital_tracking_info = form.digital_tracking_info.data
        
        db.session.commit()
        flash('Digital info updated successfully!', 'success')
        return redirect(url_for('clients.digital_info', brand_id=brand_id))
    
    elif request.method == 'GET':
        form.digital_planning_info.data = info.digital_planning_info
        form.digital_adops_info.data = info.digital_adops_info
        form.digital_tracking_info.data = info.digital_tracking_info
    
    return render_template('clients/digital_info_form.html', form=form, brand=brand, title='Edit Digital Info')

@bp.route('/brand/<int:brand_id>/digital-info/add-link', methods=['GET', 'POST'])
@login_required
def add_digital_link(brand_id):
    brand = Brand.query.get_or_404(brand_id)
    
    # Get or create digital info
    info = DigitalInfo.query.filter_by(brand_id=brand_id).first()
    if not info:
        info = DigitalInfo(brand_id=brand_id)
        db.session.add(info)
        db.session.commit()
    
    form = DigitalInfoLinkForm()
    
    if form.validate_on_submit():
        # Ensure URL has proper protocol
        url = form.url.data.strip()
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
            
        link = DigitalInfoLink(
            digital_info_id=info.id,
            link_type=form.link_type.data,
            title=form.title.data,
            url=url,
            description=form.description.data
        )
        db.session.add(link)
        db.session.commit()
        flash('Link added successfully!', 'success')
        return redirect(url_for('clients.digital_info', brand_id=brand_id))
    
    return render_template('clients/digital_link_form.html', form=form, brand=brand, title='Add Digital Link')

@bp.route('/brand/<int:brand_id>/digital-info/link/<int:link_id>/delete', methods=['POST'])
@login_required
def delete_digital_link(brand_id, link_id):
    link = DigitalInfoLink.query.get_or_404(link_id)
    
    # Verify the link belongs to this brand's digital info
    if link.digital_info.brand_id != brand_id:
        abort(404)
    
    db.session.delete(link)
    db.session.commit()
    flash('Link deleted successfully!', 'success')
    return redirect(url_for('clients.digital_info', brand_id=brand_id))